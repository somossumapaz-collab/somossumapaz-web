import mysql.connector
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import unicodedata
import re
import os

def decode_bytes(b):
    if b is None:
        return ''
    if isinstance(b, bytes):
        try:
            return b.decode('utf-8').strip()
        except UnicodeDecodeError:
            return b.decode('latin1', 'ignore').strip()
    return str(b).strip()

def normalize_str(s):
    if not s:
        return ''
    s = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('utf-8')
    return s.upper()

fem_names = {'ADELAIDA', 'ADRIANA', 'ALCIRA', 'ALEXANDRA', 'ANA', 'ANGELICA', 'ANGIE', 'BALBINA', 'BELLANIRE', 'BLANCA', 'CAROLINA', 'CATALINA', 'CLARIBEL', 'CLAUDIA', 'DANA', 'DEICY', 'DEISY', 'DEYSI', 'DIANA', 'DINA', 'DIVIA', 'DORA', 'DORIS', 'ELENA', 'ELSY', 'ENELIA', 'FLOR', 'FRANCY', 'GINNA', 'GLORIA', 'HERCILIA', 'HERMELINDA', 'JANETH', 'JASBLEYDI', 'JEIMMY', 'JENY', 'JESSICA', 'JHAZBLEIDY', 'JHURI', 'JOHANA', 'JOHANNA', 'JUANA', 'JUDITH', 'KAREN', 'KATERIN', 'LEIDY', 'LENI', 'LESLI', 'LICETH', 'LIDA', 'LIGIA', 'LILIANA', 'LINA', 'LIZETH', 'LUCELI', 'LUZ', 'MADELEY', 'MAIRA', 'MARIA', 'MARIANA', 'MARIBEL', 'MARILYN', 'MARISOL', 'MARLEN', 'MARTHA', 'MARY', 'MAYERLY', 'MELISA', 'MERY', 'MIREYA', 'MONICA', 'NANCY', 'NATALIA', 'NAYIBE', 'NEIDY', 'NICOL', 'NIDIA', 'NINFA', 'NOHORA', 'NUBIA', 'OBDULIA', 'PATRICIA', 'ROSARIO', 'SANDRA', 'SONIA', 'VIVIAN', 'VIVIANA', 'YAMILE', 'YANETH', 'YANI', 'YAQUELIN', 'YAREIDY', 'YEINY', 'YENI', 'YENIFER', 'YENNY', 'YENSI', 'YINA', 'YINETH', 'YOLIMA', 'YUDY', 'YULIANA', 'YURY'}

masc_names = {'ADRIANO', 'ALBEIRO', 'ALBINO', 'ALEJANDRO', 'ALEX', 'ALFREDO', 'ANDRES', 'AUDER', 'BRAYAN', 'CARLOS', 'CRISTIAN', 'CESAR', 'DAIRO', 'DANIEL', 'DEIVER', 'DIBER', 'DIEGO', 'DILAN', 'DUVAN', 'EDGAR', 'EMILIANO', 'ERASMO', 'FAVIO', 'FERNANDO', 'GERARDO', 'GERMAN', 'GUSTAVO', 'HARRISON', 'HEINER', 'HELMER', 'HEYDER', 'IVAN', 'JAIBER', 'JAIDER', 'JAVIER', 'JEISSON', 'JOHN', 'JOSE', 'JUAN', 'JULIO', 'KEVIN', 'LUIS', 'MAICOL', 'MANUEL', 'MAURICIO', 'MIGUEL', 'MILLER', 'NICOLAS', 'NILSON', 'NOLBERTO', 'OMAR', 'OSCAR', 'PARMENIO', 'REINALDO', 'RICARDO', 'SAMIR', 'SERGIO', 'WILLIAM', 'WILMER', 'YEISON', 'YESID', 'EDUARDO', 'OLIVERIO'}

def main():
    conn = mysql.connector.connect(
        host='srv1220.hstgr.io',
        user='u949171480_sumapaz_admin',
        password='Somossumapaz2026*',
        database='u949171480_somos_sumapaz',
        use_unicode=False
    )
    cursor = conn.cursor()

    # 1. Fetch Gender from panaca
    cursor.execute('SELECT numero_documento, genero FROM productores_panaca WHERE genero IS NOT NULL')
    genero_map = {}
    for r in cursor.fetchall():
        doc = decode_bytes(r[0])
        gen = decode_bytes(r[1])
        if doc and gen:
            genero_map[doc] = gen

    # 2. Fetch Education map
    cursor.execute('SELECT persona_id, nivel_educacion, titulo, titulo_obtenido FROM persona_educacion')
    edu_map = {}
    for r in cursor.fetchall():
        pid = decode_bytes(r[0])
        nivel = decode_bytes(r[1])
        tit = decode_bytes(r[2])
        tit_obt = decode_bytes(r[3])
        t_final = tit_obt or tit
        
        if pid not in edu_map:
            edu_map[pid] = {'niveles': [], 'titulos': []}
        if nivel and nivel not in edu_map[pid]['niveles']:
            edu_map[pid]['niveles'].append(nivel)
        if t_final and t_final not in edu_map[pid]['titulos']:
            edu_map[pid]['titulos'].append(t_final)

    # 3. Fetch personas
    cursor.execute('''
        SELECT id, nombre, tipo_documento, documento, fecha_nacimiento, 
               telefono, email, vereda, municipio_residencia, departamento_residencia,
               municipio_nacimiento, departamento_nacimiento, pais_nacimiento
        FROM persona_datos_personales
        ORDER BY id ASC
    ''')

    rows_data = []
    today = date(2026, 7, 29)

    for idx, r in enumerate(cursor.fetchall(), 1):
        pid = decode_bytes(r[0])
        nombre = decode_bytes(r[1])
        tipo_doc = decode_bytes(r[2])
        doc = decode_bytes(r[3])
        fn_str = decode_bytes(r[4])
        tel = decode_bytes(r[5])
        email = decode_bytes(r[6])
        vereda = decode_bytes(r[7]) or 'Sin especificación'
        muni_res = decode_bytes(r[8]) or 'Sumapaz'
        dept_res = decode_bytes(r[9]) or 'Bogotá D.C.'

        # Calculate Age & Birth Date format
        age = ''
        fn_formatted = ''
        if fn_str and fn_str != '0000-00-00' and len(fn_str) >= 10:
            try:
                fn = datetime.strptime(fn_str[:10], '%Y-%m-%d').date()
                fn_formatted = fn.strftime('%Y-%m-%d')
                age_val = today.year - fn.year - ((today.month, today.day) < (fn.month, fn.day))
                if 0 <= age_val <= 120:
                    age = age_val
            except Exception:
                fn_formatted = fn_str

        # Resolve Gender
        gen = genero_map.get(doc, '')
        if not gen:
            norm_n = normalize_str(nombre)
            words = norm_n.split()
            gen = 'No especificado'
            for w in words:
                if w in fem_names or 'MARIA' in w:
                    gen = 'Femenino'
                    break
                elif w in masc_names:
                    gen = 'Masculino'
                    break

        # Education
        ed_info = edu_map.get(pid, {'niveles': [], 'titulos': []})
        nivel_edu = ', '.join(ed_info['niveles']) if ed_info['niveles'] else 'Sin registro'
        titulo_edu = ', '.join(ed_info['titulos']) if ed_info['titulos'] else 'N/A'

        rows_data.append({
            'num': idx,
            'id': pid,
            'nombre': nombre,
            'tipo_doc': tipo_doc,
            'doc': doc,
            'fecha_nacimiento': fn_formatted,
            'edad': age,
            'sexo': gen,
            'nivel_educativo': nivel_edu,
            'titulo_obtenido': titulo_edu,
            'vereda': vereda,
            'muni_residencia': muni_res,
            'dept_residencia': dept_res,
            'telefono': tel,
            'email': email
        })

    conn.close()

    # Create OpenPyXL Workbook
    wb = openpyxl.Workbook()
    
    # Styles
    font_title = Font(name='Segoe UI', size=16, bold=True, color='1F4E79')
    font_subtitle = Font(name='Segoe UI', size=11, italic=True, color='595959')
    font_section = Font(name='Segoe UI', size=13, bold=True, color='1F4E79')
    font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    font_data = Font(name='Segoe UI', size=10, color='333333')
    
    fill_header = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    fill_zebra = PatternFill(start_color='F2F4F7', end_color='F2F4F7', fill_type='solid')
    fill_card = PatternFill(start_color='E9EEF4', end_color='E9EEF4', fill_type='solid')
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_side = Side(border_style='thin', color='D3D3D3')
    border_data = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_card = Border(left=Side(style='medium', color='1F4E79'), right=thin_side, top=thin_side, bottom=thin_side)

    # ----------------------------------------------------
    # SHEET 1: Data Table ("Hojas de Vida - Demografía")
    # ----------------------------------------------------
    ws_data = wb.active
    ws_data.title = "Hojas de Vida - Demografía"
    ws_data.views.sheetView[0].showGridLines = True

    # Header Title Block
    ws_data.merge_cells("A1:O1")
    ws_data["A1"] = "REPORTE DE DATOS DEMOGRÁFICOS - HOJAS DE VIDA"
    ws_data["A1"].font = font_title
    ws_data["A1"].alignment = Alignment(horizontal='left', vertical='center')

    ws_data.merge_cells("A2:O2")
    ws_data["A2"] = f"Alcaldía Local de Sumapaz | Generado el {datetime.now().strftime('%Y-%m-%d')} | Total Registros: {len(rows_data)}"
    ws_data["A2"].font = font_subtitle
    ws_data["A2"].alignment = Alignment(horizontal='left', vertical='center')

    ws_data.row_dimensions[1].height = 28
    ws_data.row_dimensions[2].height = 20
    ws_data.row_dimensions[3].height = 10  # blank space

    headers = [
        "N°", "ID", "Nombre Completo", "Tipo Doc.", "Número Documento",
        "Fecha Nacimiento", "Edad", "Sexo / Género", "Nivel Educativo",
        "Título / Formación Obtenida", "Vereda", "Municipio Residencia",
        "Departamento Residencia", "Teléfono", "Correo Electrónico"
    ]

    header_row = 4
    ws_data.row_dimensions[header_row].height = 28

    for col_num, h in enumerate(headers, 1):
        cell = ws_data.cell(row=header_row, column=col_num, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_data

    start_row = 5
    for idx, d in enumerate(rows_data, start_row):
        row_cells = [
            ws_data.cell(row=idx, column=1, value=d['num']),
            ws_data.cell(row=idx, column=2, value=int(d['id']) if d['id'].isdigit() else d['id']),
            ws_data.cell(row=idx, column=3, value=d['nombre']),
            ws_data.cell(row=idx, column=4, value=d['tipo_doc']),
            ws_data.cell(row=idx, column=5, value=d['doc']),
            ws_data.cell(row=idx, column=6, value=d['fecha_nacimiento']),
            ws_data.cell(row=idx, column=7, value=d['edad']),
            ws_data.cell(row=idx, column=8, value=d['sexo']),
            ws_data.cell(row=idx, column=9, value=d['nivel_educativo']),
            ws_data.cell(row=idx, column=10, value=d['titulo_obtenido']),
            ws_data.cell(row=idx, column=11, value=d['vereda']),
            ws_data.cell(row=idx, column=12, value=d['muni_residencia']),
            ws_data.cell(row=idx, column=13, value=d['dept_residencia']),
            ws_data.cell(row=idx, column=14, value=d['telefono']),
            ws_data.cell(row=idx, column=15, value=d['email'])
        ]

        ws_data.row_dimensions[idx].height = 22
        is_even = (idx % 2 == 0)

        for col_idx, cell in enumerate(row_cells, 1):
            cell.font = font_data
            cell.border = border_data
            if is_even:
                cell.fill = fill_zebra
            
            # Alignments
            if col_idx in [1, 2, 4, 5, 6, 7, 8, 14]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # Column widths
    ws_data.column_dimensions['A'].width = 8   # N°
    ws_data.column_dimensions['B'].width = 8   # ID
    ws_data.column_dimensions['C'].width = 36  # Nombre Completo
    ws_data.column_dimensions['D'].width = 24  # Tipo Doc
    ws_data.column_dimensions['E'].width = 20  # Documento
    ws_data.column_dimensions['F'].width = 18  # Fecha Nacimiento
    ws_data.column_dimensions['G'].width = 10  # Edad
    ws_data.column_dimensions['H'].width = 16  # Sexo
    ws_data.column_dimensions['I'].width = 30  # Nivel Educativo
    ws_data.column_dimensions['J'].width = 40  # Título / Formación
    ws_data.column_dimensions['K'].width = 24  # Vereda
    ws_data.column_dimensions['L'].width = 22  # Muni Res
    ws_data.column_dimensions['M'].width = 22  # Dept Res
    ws_data.column_dimensions['N'].width = 18  # Teléfono
    ws_data.column_dimensions['O'].width = 34  # Email

    # ----------------------------------------------------
    # SHEET 2: Summary Dashboard ("Resumen Demográfico")
    # ----------------------------------------------------
    ws_sum = wb.create_sheet(title="Resumen Demográfico")
    ws_sum.views.sheetView[0].showGridLines = True

    ws_sum.merge_cells("A1:G1")
    ws_sum["A1"] = "DASHBOARD DEMOGRÁFICO Y ESTADÍSTICAS"
    ws_sum["A1"].font = font_title
    ws_sum["A1"].alignment = Alignment(horizontal='left', vertical='center')

    ws_sum.row_dimensions[1].height = 28
    ws_sum.row_dimensions[2].height = 10

    # Calculate statistics
    total_persons = len(rows_data)
    ages_list = [d['edad'] for d in rows_data if isinstance(d['edad'], int)]
    avg_age = round(sum(ages_list)/len(ages_list), 1) if ages_list else 0

    gender_counts = {'Femenino': 0, 'Masculino': 0, 'No especificado': 0}
    vereda_counts = {}
    edu_counts = {}
    age_groups = {'< 18 años': 0, '18 - 29 años (Jóvenes)': 0, '30 - 49 años': 0, '50 - 64 años': 0, '65+ años': 0}

    for d in rows_data:
        g = d['sexo']
        gender_counts[g] = gender_counts.get(g, 0) + 1
        
        v = d['vereda']
        vereda_counts[v] = vereda_counts.get(v, 0) + 1
        
        ne = d['nivel_educativo']
        edu_counts[ne] = edu_counts.get(ne, 0) + 1

        a = d['edad']
        if isinstance(a, int):
            if a < 18:
                age_groups['< 18 años'] += 1
            elif 18 <= a <= 29:
                age_groups['18 - 29 años (Jóvenes)'] += 1
            elif 30 <= a <= 49:
                age_groups['30 - 49 años'] += 1
            elif 50 <= a <= 64:
                age_groups['50 - 64 años'] += 1
            else:
                age_groups['65+ años'] += 1

    # KPI Summary Cards (Rows 3 to 6)
    kpis = [
        ("Total Hojas de Vida", total_persons, "Personas registradas"),
        ("Promedio de Edad", f"{avg_age} años", "Rango etario de la población"),
        ("Mujeres", f"{gender_counts['Femenino']} ({round(gender_counts['Femenino']/total_persons*100, 1)}%)", "Población femenina"),
        ("Hombres", f"{gender_counts['Masculino']} ({round(gender_counts['Masculino']/total_persons*100, 1)}%)", "Población masculina"),
        ("Veredas Cubiertas", len(vereda_counts), "Sectores de Sumapaz")
    ]

    col_positions = [1, 3, 5, 7, 9]
    for idx, (title, val, sub) in enumerate(kpis):
        start_c = col_positions[idx]
        end_c = start_c + 1
        
        ws_sum.merge_cells(start_row=3, start_column=start_c, end_row=3, end_column=end_c)
        ws_sum.merge_cells(start_row=4, start_column=start_c, end_row=4, end_column=end_c)
        ws_sum.merge_cells(start_row=5, start_column=start_c, end_row=5, end_column=end_c)

        cell_title = ws_sum.cell(row=3, column=start_c, value=title)
        cell_val = ws_sum.cell(row=4, column=start_c, value=val)
        cell_sub = ws_sum.cell(row=5, column=start_c, value=sub)

        cell_title.font = Font(name='Segoe UI', size=9, bold=True, color='595959')
        cell_val.font = Font(name='Segoe UI', size=16, bold=True, color='1F4E79')
        cell_sub.font = Font(name='Segoe UI', size=8, italic=True, color='7F7F7F')

        cell_title.alignment = align_center
        cell_val.alignment = align_center
        cell_sub.alignment = align_center

        for r in range(3, 6):
            for c in range(start_c, end_c + 1):
                cell_box = ws_sum.cell(row=r, column=c)
                cell_box.fill = fill_card
                cell_box.border = border_card

    # Section 1: Distribución por Vereda (Columns A & B & C)
    ws_sum.cell(row=8, column=1, value="Distribución por Vereda").font = font_section
    ws_sum.cell(row=9, column=1, value="Vereda").font = font_header
    ws_sum.cell(row=9, column=1).fill = fill_header
    ws_sum.cell(row=9, column=1).alignment = align_center

    ws_sum.cell(row=9, column=2, value="Cantidad").font = font_header
    ws_sum.cell(row=9, column=2).fill = fill_header
    ws_sum.cell(row=9, column=2).alignment = align_center

    ws_sum.cell(row=9, column=3, value="%").font = font_header
    ws_sum.cell(row=9, column=3).fill = fill_header
    ws_sum.cell(row=9, column=3).alignment = align_center

    sorted_veredas = sorted(vereda_counts.items(), key=lambda x: x[1], reverse=True)
    curr_row = 10
    for v_name, count in sorted_veredas:
        pct = round(count / total_persons * 100, 1)
        c1 = ws_sum.cell(row=curr_row, column=1, value=v_name)
        c2 = ws_sum.cell(row=curr_row, column=2, value=count)
        c3 = ws_sum.cell(row=curr_row, column=3, value=f"{pct}%")
        
        c1.font = font_data
        c2.font = font_data
        c3.font = font_data
        c1.border = border_data
        c2.border = border_data
        c3.border = border_data
        c1.alignment = align_left
        c2.alignment = align_center
        c3.alignment = align_center
        curr_row += 1

    # Section 2: Distribución por Nivel Educativo (Columns E & F & G)
    ws_sum.cell(row=8, column=5, value="Distribución por Nivel Educativo").font = font_section
    ws_sum.cell(row=9, column=5, value="Nivel Educativo").font = font_header
    ws_sum.cell(row=9, column=5).fill = fill_header
    ws_sum.cell(row=9, column=5).alignment = align_center

    ws_sum.cell(row=9, column=6, value="Cantidad").font = font_header
    ws_sum.cell(row=9, column=6).fill = fill_header
    ws_sum.cell(row=9, column=6).alignment = align_center

    ws_sum.cell(row=9, column=7, value="%").font = font_header
    ws_sum.cell(row=9, column=7).fill = fill_header
    ws_sum.cell(row=9, column=7).alignment = align_center

    sorted_edu = sorted(edu_counts.items(), key=lambda x: x[1], reverse=True)
    curr_row_edu = 10
    for e_name, count in sorted_edu:
        pct = round(count / total_persons * 100, 1)
        c1 = ws_sum.cell(row=curr_row_edu, column=5, value=e_name)
        c2 = ws_sum.cell(row=curr_row_edu, column=6, value=count)
        c3 = ws_sum.cell(row=curr_row_edu, column=7, value=f"{pct}%")
        
        c1.font = font_data
        c2.font = font_data
        c3.font = font_data
        c1.border = border_data
        c2.border = border_data
        c3.border = border_data
        c1.alignment = align_left
        c2.alignment = align_center
        c3.alignment = align_center
        curr_row_edu += 1

    # Section 3: Distribución por Grupos de Edad (Columns I & J & K)
    ws_sum.cell(row=8, column=9, value="Distribución por Rango Etario").font = font_section
    ws_sum.cell(row=9, column=9, value="Rango Etario").font = font_header
    ws_sum.cell(row=9, column=9).fill = fill_header
    ws_sum.cell(row=9, column=9).alignment = align_center

    ws_sum.cell(row=9, column=10, value="Cantidad").font = font_header
    ws_sum.cell(row=9, column=10).fill = fill_header
    ws_sum.cell(row=9, column=10).alignment = align_center

    ws_sum.cell(row=9, column=11, value="%").font = font_header
    ws_sum.cell(row=9, column=11).fill = fill_header
    ws_sum.cell(row=9, column=11).alignment = align_center

    curr_row_age = 10
    for range_name, count in age_groups.items():
        pct = round(count / total_persons * 100, 1)
        c1 = ws_sum.cell(row=curr_row_age, column=9, value=range_name)
        c2 = ws_sum.cell(row=curr_row_age, column=10, value=count)
        c3 = ws_sum.cell(row=curr_row_age, column=11, value=f"{pct}%")
        
        c1.font = font_data
        c2.font = font_data
        c3.font = font_data
        c1.border = border_data
        c2.border = border_data
        c3.border = border_data
        c1.alignment = align_left
        c2.alignment = align_center
        c3.alignment = align_center
        curr_row_age += 1

    # Adjust widths for summary sheet
    ws_sum.column_dimensions['A'].width = 28
    ws_sum.column_dimensions['B'].width = 14
    ws_sum.column_dimensions['C'].width = 12
    ws_sum.column_dimensions['D'].width = 6
    ws_sum.column_dimensions['E'].width = 38
    ws_sum.column_dimensions['F'].width = 14
    ws_sum.column_dimensions['G'].width = 12
    ws_sum.column_dimensions['H'].width = 6
    ws_sum.column_dimensions['I'].width = 28
    ws_sum.column_dimensions['J'].width = 14
    ws_sum.column_dimensions['K'].width = 12

    # Save output file
    output_filename = "Datos_Demograficos_Hojas_de_Vida.xlsx"
    wb.save(output_filename)
    print(f"¡Éxito! Archivo guardado correctamente como: {output_filename}")

if __name__ == '__main__':
    main()
